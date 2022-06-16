# !/usr/bin python3                                 
# encoding: utf-8 -*-   
# @file     : file_operate.py                       
# @author   : 沙陌 Matongxue_2
# @Time     : 2022-05-04 9:42
# @Copyright: 北京码同学
import openpyxl
import yaml

def read_excel(filepath,sheet_name):
    wb =openpyxl.load_workbook(filepath) #得到整个excel文档对象
    sheet_data = wb[sheet_name] # 获取某个sheet工作表的数据
    data = [] # 定义一个列表，用来存储多行数据，每行数据都是一个列表
    # 遍历sheet_data，整合各个单元格的数据，让其达到目标数据结构
    lines_count = sheet_data.max_row # 得到总行数
    cols_count = sheet_data.max_column # 得到总列数
    # 对于sheet_data来说他的行和列的索引顺序都是从1开始的
    for l in range(2,lines_count+1):
        line = [] # 定一个列表，用来存储当前行的每个单元格数据
        for c in range(1,cols_count+1):
            cell = sheet_data.cell(l,c).value # 获取单元格数据
            if cell is None: #如果单元格为空，给他赋值空字符串
                cell = ''
            line.append(cell)
        data.append(line) # 将当前行列表存储到最终数据data中
    return data

def write_excel(filepath,sheet_name,row,col,text):
    wb = openpyxl.load_workbook(filepath)
    sheet_data = wb[sheet_name]
    sheet_data.cell(row=row,column=col,value=text) #写入数据
    wb.save(filepath) # 保存


def read_yaml(filepath):
    with open(filepath,mode='r',encoding='UTF-8') as f:
        content = yaml.load(f,Loader=yaml.FullLoader)
        return content

def write_yaml(filepath,content):
    with open(filepath,mode='w',encoding='UTF-8') as f:
        yaml.dump(content,f,Dumper=yaml.Dumper)
if __name__ == '__main__':
    # print(read_excel('test_data.xlsx', '登录接口数据'))

    # 写文件只是测试一下，写完之后记得清除掉写入的数据，否则会影响后续的数据读取
    # write_excel('test_data.xlsx', '登录接口数据',8,9,'测试一下')
    # print(read_yaml('test_data.yml')['登录接口数据'])
    write_yaml('test_data1.yaml',{'user':'shamo','pwd':'123456'})

